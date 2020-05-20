import csv
import json
import dicttoxml
from xml.dom.minidom import parseString
from pyduq.duqerror import ValidationError
from unidecode import unidecode
from openpyxl import Workbook,load_workbook

class FileTools(object):
    """ FileTools: 
    This is a utility class to help manage data files.
    The file is expected to be in a CSV format with a header row.
    Note: This could be replaced by Pandas
    """

    @staticmethod
    def csvFileToDict(fileName:str) -> dict:
        """ csvFileToDict:
        Converts a csv file into a dictionary of dictionaries.
        Each row is its own dictionary with each attribute recorded as a tupple,
        indexed by a row counter.
        """
        data = dict()
        resultset = []
        
        # first we load the data into a simple 
        with open(fileName, errors='ignore') as f:
            reader = csv.DictReader(f)
                               
            # Get a list of the column names returned from the file
            columns = list(reader.fieldnames)
            
            for row in reader:
                resultset.append(row)
                
            for col in columns:
                data[col]= [("(Null)" if row[col] is None else FileTools.FormatString(str(row[col]).strip())) for row in resultset]            
        
        return data


    @staticmethod
    def xlsFileToDict(fileName:str) -> dict:
        """ xlsFileToDict:
        Converts an Excel spreadsheet into a dictionary of dictionaries.
        Each row is its own dictionary with each attribute recorded as a tupple,
        indexed by a row counter.
        
        Assumptions: The spreadsheet is well-formatted columns and rows.
        """
        data = {}
        
        # first we load the data into a simple
        workbook = load_workbook(filename=fileName, data_only=True, read_only = True)
        sheet = workbook.active
        
        # extract the column headers - assumes headers in row 1 only - perhaps 
        # this could be configrable :-)
        columns = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

        # convert the data from rows into columns. This looks clunky
        # but it's somehow faster that iterating through the columns and
        # and provides an oportunity to clean up indiuvidual cell values.
        for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
            col=0
            
            for value in row:
                if (not value is None):
                    if (not columns[col] in data):
                        data[columns[col]] = []
                        
                    data[columns[col]].append("(Null)" if value is None else FileTools.FormatString(str(value).strip()))            
                    col+=1
        
        del(sheet)
        del(workbook)
        return data

  
    @staticmethod
    def JSONtoMeta(fileName:str) ->dict:
        meta = {}
        
        with open(fileName, 'r', errors='ignore') as json_file:
            meta = json.load(json_file)
          
        return meta
  
    @staticmethod
    def MetatoXMLFile(xml_filename:str, meta:dict):
        xml = dicttoxml.dicttoxml(meta)
        dom = parseString(xml)

        with open(xml_filename, 'w') as w:
            w.write(dom.toprettyxml())
            
    @staticmethod
    def MetatoJSONFile(JSON_filename:str, meta:dict):
        json_dict = json.dumps(meta, indent=4)
        
        with open(JSON_filename, 'w') as w:
            w.write(json_dict)
            
    
    @staticmethod
    def FormatString(s:str) ->str:
        # converts a unicode string to ascii
        
        if isinstance(s, str):
            try:
                s.encode('ascii')
                return s
            except:
                return unidecode(s)
        else:
          return s


    @staticmethod
    def inferMeta(dataset:dict) ->dict:
        meta = {}
    
        print("Infering the metadata...","\r")

        for attribute_key, attributes in dataset.items():
            print("Analysing the data...\t" + attribute_key)

            metarow = {}
            
            isMandatory = True
            isInt = True
            isFloat = True
            isBool = True
            isDate = False
            
            seen=set()
            size=0

                        
            for value in attributes:                
                #check to see if there are any blanks or nulls in order to determine if values are mandatory
                if (len(value)==0 or value == "(Null)"):
                    isMandatory = False
                else: 
                    #this will help determine if there are any duplicates                
                    if (not value in seen):
                        seen.add(value)

                    #calculate the size of the column
                    if (len(value)>size):
                        size = len(value)
                    
                    #test for integer
                    try:
                        int(value)
                    except Exception as e:
                        isInt = False

                    #test for float
                    try:
                        float(value)
                    except Exception as e:
                        isFloat = False
                    
                    if (not value.lower() in ["0","1","no","yes","false","true"]):
                        isBool = False
            
                    if (attribute_key.lower().find("date") != -1):
                        isDate = True
            
            metarow["Size"] = size
            metarow["Mandatory"] = isMandatory
            metarow["AllowBlank"] = not isMandatory
            metarow["Type"] = ("integer" if isInt else "float" if isFloat else "boolean" if isBool else "date" if isDate else "string")
            metarow["Unique"] = (len(seen) == len(attributes))

            #if we have a small number of items in our seen set then lets presume this is an enumerated attributed
            #note - there is no minimum number of set elements - an alternative approach is to capture the top n repeating strings -
            #but you're likely to get the same outcome
            if (len(seen)< 100):
                metarow["Enum"] = str(sorted(list(seen)))
            
            meta[attribute_key]=metarow
            
        return meta
    
    @staticmethod
    def saveProfile(outputFile, data_profile:list):
        if (len(data_profile)>0):
            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet()
            c=data_profile[0]
            headers = list(c.keys())
            sheet.append(headers)
            
            for x in data_profile:
                sheet.append(list(x.values()))
        
            workbook.save(filename=outputFile)
            workbook.close()
            del(sheet)
            del(workbook)
