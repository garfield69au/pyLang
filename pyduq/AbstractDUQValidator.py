import abc
from openpyxl import Workbook
from pyduq.dataqualityerror import DataQualityError, DataQualityDimension
from pyduq.dataprofile import DataProfile
from pyduq.duqerror import ValidationError
   
class AbstractDUQValidator(abc.ABC):
    """ AbstractDQValidator: 
    Base class that any data quality validation class should implement. It provides some basic structure for a validator.
    The constructor expects a resultset dictionary and a metadata disctionary. Classes that implement this base class will have access to
    a copy of both of those objects.
    The cobstructor will also create an empty list of errors and a empty list of measurement counters.
    """

    def __init__(self:object, dataset:dict, meta:dict):
        
        if (dataset is None):
            raise ValidationError("LANG Exception: DataSet has not been set", None)
        
        if (meta is None):
            raise ValidationError("LANG Exception: metadata has not been set", None)
        
        self.metadata = meta
        self.dataset = dataset
        self.validation_errors = []
        self.data_profile = []
            
        
    def addDataQualityError(self:object, data_quality_error:DataQualityError):
        """
        Add a new dimension.
        """
        if (data_quality_error is None):
            raise ValidationError("LANG Exception: DataQualityError has not been set", None)
        
        self.validation_errors.append(data_quality_error.to_dict())
        
    
    def profileData(self:object, meta_attribute_definition:dict, colData:dict, key:str):
        if (colData is None):
            raise ValidationError("LANG Exception: Coldata has not been set", None)
        
        profile = DataProfile()
        profile.profileData(meta_attribute_definition, colData, key)
        profile.setPosition(len(self.data_profile)+1)
        self.data_profile.append(profile.to_dict())
        

    def saveCountersSummary(self:object, outputFile):
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet()
        headers = ["attribute"]
        
        for name in DataQualityDimension.names():
            headers.append(name)
            headers.append(name + " SCORE")
            
        sheet.append(headers)

        summary = self.summariseCounters()
    
        for datarow in summary.values():        
            for row in datarow:
                sheet.append(list(row.values()))
            
        workbook.save(filename=outputFile)
        workbook.close()
        del(sheet)
        del(workbook)


    def summariseCounters(self:object) ->dict:
        
        # get the MeasurementCategory Enum as a list
        categories = list(DataQualityDimension)
        summary = {}
        attribute_errors = {}
 
        # Construct a list of errors per attribute and store in a dict 
        for item in self.validation_errors:
            key = item["attribute"]
            if (not key in attribute_errors):
                attribute_errors[key] = []
            
            attribute_errors[key].append(item)
        
        # now count how many times each category appears for each attribute
        #for item, data in attribute_errors.items():
        for item in self.metadata:
            summary_row = dict()
            summary_row["attribute"] = item

            summary[item]=[]
            data = dict()
            
            if (item in attribute_errors):
                data = attribute_errors[item]
            
            for name in categories:
                error_count=0
                
                for d in data:
                    if (d["error_dimension"] == name.value):
                        error_count+=1
                
                # for each attribute create a list of dictionaries contaning a count of each category
                summary_row[name.name] = error_count
                    
                try:
                    
                    score = error_count / len(self.dataset[item])
                    summary_row[name.name + " SCORE"] = round(score, 6)
                   
                except Exception as e:
                        summary_row[name.name + " SCORE"] =  0
                        
            summary[item].append(summary_row)
                                
        return summary
        

    def saveCounters(self:object, outputFile):
        if (len(self.validation_errors)>0):
            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet()
            headers = list(self.validation_errors[0].keys())
            sheet.append(headers)
        
            for y in self.validation_errors:
                sheet.append(list(y.values()))
        
            workbook.save(filename=outputFile)
            workbook.close()
            del(sheet)
            del(workbook)

    
    @abc.abstractmethod
    def validate(self:object):
        pass

        
    @abc.abstractmethod
    def validateList(self:object, key:str):
        pass        

